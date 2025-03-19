import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
# Laad het bestand met de gemarkeerde cellen
def get_colored_cells(file_path):
    # Open het Excel-bestand
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active  # Actieve sheet

    colored_cells = []

    # Loop door alle cellen in het werkblad
    for row in sheet.iter_rows():
        for cell in row:
            # Controleer of de cel een achtergrondkleur heeft (gevuld)
            if cell.fill.patternType == 'solid' and cell.fill.start_color.index != '00000000':  # Als de cel gekleurd is
                color = cell.fill.start_color.index  # Haal de kleurcode op
                colored_cells.append((cell.row, cell.column, cell.value, color))  # Bewaar rij, kolom, waarde, en kleur
    return colored_cells

# Laad het bestand waarin je de tekst wilt zoeken en kleuren
def apply_colors_to_csv(excel_file, csv_file):
    # Haal de gekleurde cellen op uit het Excel-bestand
    colored_cells = get_colored_cells(excel_file)

    # Laad het CSV-bestand
    df = pd.read_csv(csv_file)

    # Voeg een nieuwe kolom toe voor de kleurcode
    df['color'] = None

    
    for _, _, text, color in colored_cells:
        
        df.loc[df['Text_mutatie'] == text, 'color'] = color 
        df.loc[df['Text_mutatie'] == text, 'color'] = color

    # Sla het CSV-bestand op met de kleurcode
    df.to_csv('output_with_colors.csv', index=False)
    print("Kleurcodes toegevoegd en bestand opgeslagen als 'output_with_colors.csv'.")

# Gebruik de functie
source_file = 'mail output labels.xlsx'  
target_file = '/workspaces/nlptest/all_mails (1).csv'  
apply_colors_to_csv(source_file, target_file)
